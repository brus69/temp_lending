import csv
from decimal import Decimal, InvalidOperation

from django import forms
from django.contrib import admin
from django.contrib import messages
from django.contrib.admin.helpers import ActionForm
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.html import format_html

from openpyxl import Workbook, load_workbook

from .models import (
    Article,
    Category,
    CategorySpecAttribute,
    Favorite,
    News,
    Order,
    OrderItem,
    Organization,
    Product,
    ProductLabel,
    ProductQuestion,
    ProductReview,
    ProductReviewPhoto,
    ProductGalleryImage,
    ProductSpecValue,
    Promotion,
    UserProfile,
)


class CategorySpecAttributeInline(admin.TabularInline):
    model = CategorySpecAttribute
    extra = 1
    ordering = ("sort_order", "id")


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = (
        "name",
        "slug",
        "show_in_top_menu",
        "default_listing_view",
        "product_count",
        "is_featured",
        "sort_order",
    )
    list_editable = ("show_in_top_menu", "is_featured", "sort_order")
    search_fields = ("name", "slug")
    inlines = [CategorySpecAttributeInline]
    readonly_fields = ("product_count", "image_preview")

    @admin.display(description="Предпросмотр изображения")
    def image_preview(self, obj: Category) -> str:
        url = obj.primary_image_url if obj and obj.pk else ""
        if not url:
            return "—"
        return format_html('<img src="{}" alt="" style="max-height:140px;max-width:100%;object-fit:contain" />', url)

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "name",
                    "slug",
                    "image",
                    "image_url",
                    "image_preview",
                    "is_featured",
                    "show_in_top_menu",
                    "sort_order",
                    "default_listing_view",
                ),
            },
        ),
        ("SEO", {"fields": ("meta_title", "meta_description"), "classes": ("collapse",)}),
        (
            "Каталог",
            {
                "fields": ("product_count",),
                "description": "Считается автоматически по товарам в категории.",
            },
        ),
    )


class ProductReviewPhotoInline(admin.TabularInline):
    model = ProductReviewPhoto
    extra = 0


@admin.register(ProductReview)
class ProductReviewAdmin(admin.ModelAdmin):
    list_display = ("product", "user", "rating", "created_at")
    list_filter = ("rating", "created_at")
    search_fields = ("product__name", "user__email", "text")
    inlines = [ProductReviewPhotoInline]


@admin.register(ProductQuestion)
class ProductQuestionAdmin(admin.ModelAdmin):
    list_display = ("product", "user", "created_at", "answered_at")
    list_filter = ("created_at", "answered_at")
    search_fields = ("product__name", "text", "answer_text")


class ProductSpecValueInline(admin.TabularInline):
    model = ProductSpecValue
    extra = 1

    def get_formset(self, request, obj=None, **kwargs):
        category_id = obj.category_id if obj else None

        class ProductSpecValueAdminForm(forms.ModelForm):
            class Meta:
                model = ProductSpecValue
                fields = "__all__"

            def __init__(self, *args, **form_kw):
                super().__init__(*args, **form_kw)
                qs = (
                    CategorySpecAttribute.objects.filter(category_id=category_id).order_by("sort_order", "id")
                    if category_id
                    else CategorySpecAttribute.objects.none()
                )
                self.fields["attribute"].queryset = qs

        kwargs.setdefault("form", ProductSpecValueAdminForm)
        return super().get_formset(request, obj, **kwargs)


class ProductGalleryImageInline(admin.TabularInline):
    model = ProductGalleryImage
    extra = 1
    max_num = ProductGalleryImage.MAX_IMAGES_PER_PRODUCT
    validate_max = True
    fields = ("image", "sort_order")
    ordering = ("sort_order", "id")


PRODUCT_IMPORT_EXPORT_FIELDS = [
    "category_slug",
    "name",
    "slug",
    "sku",
    "price",
    "old_price",
    "image_url",
    "description",
    "stock_store",
    "stock_warehouse",
    "labels",
]


class ProductImportForm(forms.Form):
    file = forms.FileField(label="Файл")
    use_csv = forms.BooleanField(label="CSV (снять для Excel)", required=False, initial=False)


class ProductExportActionForm(ActionForm):
    use_csv = forms.BooleanField(label="CSV (по умолчанию Excel)", required=False, initial=False)


class ProductAdminForm(forms.ModelForm):
    class Meta:
        model = Product
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        redirect_qs = Product.objects.filter(is_active=True).order_by("name")
        if self.instance.pk:
            redirect_qs = redirect_qs.exclude(pk=self.instance.pk)
        self.fields["redirect_product"].queryset = redirect_qs
        self.fields["redirect_product"].help_text = (
            "Укажите активный товар — с неактивной карточки будет ответ 301."
        )


@admin.register(ProductLabel)
class ProductLabelAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "background_class", "sort_order")
    list_editable = ("sort_order",)
    search_fields = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}
    ordering = ("sort_order", "name")
    fieldsets = (
        (None, {"fields": ("name", "slug", "sort_order")}),
        (
            "Оформление",
            {
                "fields": ("background_class", "text_class", "detail_banner_text"),
                "description": "Классы Tailwind для бейджа на витрине. Текст на странице товара — необязательный блок под галереей.",
            },
        ),
    )


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    form = ProductAdminForm
    list_display = ("name", "sku", "category", "price", "is_active", "labels_display")
    list_filter = ("is_active", "category", "labels")
    search_fields = ("name", "sku", "slug")
    autocomplete_fields = ("redirect_product",)
    filter_horizontal = ("labels",)
    inlines = [ProductSpecValueInline, ProductGalleryImageInline]
    readonly_fields = ("rating", "reviews_count", "image_preview")
    actions = ("export_products",)
    action_form = ProductExportActionForm
    change_list_template = "admin/shop/product/change_list.html"
    change_form_template = "admin/shop/product/change_form.html"

    class Media:
        css = {"all": ("shop/admin/product_status.css",)}
        js = ("shop/admin/product_status.js",)

    @admin.display(description="Метки")
    def labels_display(self, obj: Product) -> str:
        if not obj.pk:
            return "—"
        return ", ".join(obj.labels.order_by("sort_order", "name").values_list("name", flat=True)) or "—"

    @admin.display(description="Предпросмотр изображения")
    def image_preview(self, obj: Product) -> str:
        url = obj.primary_image_url if obj and obj.pk else ""
        if not url:
            return "—"
        return format_html('<img src="{}" alt="" style="max-height:140px;max-width:100%;object-fit:contain" />', url)

    fieldsets = (
        (
            None,
            {
                "fields": (
                    "category",
                    "name",
                    "slug",
                    "sku",
                    "price",
                    "old_price",
                    "image",
                    "image_url",
                    "image_preview",
                    "description",
                ),
            },
        ),
        ("SEO", {"fields": ("meta_title", "meta_description"), "classes": ("collapse",)}),
        (
            "Публикация",
            {
                "fields": ("is_active", "redirect_product"),
                "description": "Снимите «Товар активен», чтобы скрыть товар с витрины и настроить редирект 301.",
            },
        ),
        (
            "Остатки и метки",
            {
                "fields": (
                    "stock_store",
                    "stock_warehouse",
                    "labels",
                ),
            },
        ),
        ("Рейтинг (авто)", {"fields": ("rating", "reviews_count")}),
    )

    def save_model(self, request, obj, form, change):
        prev_cat = None
        if change and obj.pk:
            prev_cat = Product.objects.filter(pk=obj.pk).values_list("category_id", flat=True).first()
        super().save_model(request, obj, form, change)
        if change and prev_cat is not None and prev_cat != obj.category_id:
            ProductSpecValue.objects.filter(product=obj).exclude(attribute__category_id=obj.category_id).delete()
        Category.sync_product_counts()

    def delete_model(self, request, obj):
        super().delete_model(request, obj)
        Category.sync_product_counts()

    def delete_queryset(self, request, queryset):
        super().delete_queryset(request, queryset)
        Category.sync_product_counts()

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-products/",
                self.admin_site.admin_view(self.import_products_view),
                name="shop_product_import",
            ),
        ]
        return custom_urls + urls

    @admin.action(description="Экспорт выбранных товаров")
    def export_products(self, request: HttpRequest, queryset):
        as_csv = request.POST.get("use_csv") in {"on", "true", "1"}
        if as_csv:
            return self._export_products_csv(queryset)
        return self._export_products_xlsx(queryset)

    def _export_products_csv(self, queryset) -> HttpResponse:
        headers, rows = self._build_export_payload(queryset)
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="products_export.csv"'
        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        return response

    def _export_products_xlsx(self, queryset) -> HttpResponse:
        headers, rows = self._build_export_payload(queryset)
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(headers)
        for row in rows:
            ws.append(row)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="products_export.xlsx"'
        wb.save(response)
        return response

    def _build_export_payload(self, queryset):
        products = list(
            queryset.select_related("category").prefetch_related(
                "spec_values__attribute",
                "labels",
            ),
        )
        spec_headers: list[str] = []
        spec_header_set: set[str] = set()
        product_specs: dict[int, dict[str, str]] = {}

        for product in products:
            spec_map: dict[str, str] = {}
            ordered_specs = sorted(
                product.spec_values.all(),
                key=lambda sv: (sv.attribute.sort_order, sv.attribute_id),
            )
            for sv in ordered_specs:
                header = f"spec__{sv.attribute.name}"
                if header not in spec_header_set:
                    spec_header_set.add(header)
                    spec_headers.append(header)
                spec_map[header] = sv.value
            product_specs[product.id] = spec_map

        headers = [*PRODUCT_IMPORT_EXPORT_FIELDS, *spec_headers]
        rows = []
        for product in products:
            base_row = [
                product.category.slug,
                product.name,
                product.slug,
                product.sku,
                str(product.price),
                str(product.old_price) if product.old_price is not None else "",
                product.image_url,
                product.description,
                product.stock_store,
                product.stock_warehouse,
                ";".join(
                    product.labels.order_by("sort_order", "name").values_list("slug", flat=True),
                ),
            ]
            spec_map = product_specs.get(product.id, {})
            spec_row = [spec_map.get(header, "") for header in spec_headers]
            rows.append([*base_row, *spec_row])
        return headers, rows

    def import_products_view(self, request: HttpRequest):
        if request.method == "POST":
            form = ProductImportForm(request.POST, request.FILES)
            if form.is_valid():
                imported_count, updated_count, errors = self._import_products_from_file(
                    upload=form.cleaned_data["file"],
                    as_csv=form.cleaned_data["use_csv"],
                )
                for err in errors[:10]:
                    self.message_user(request, err, level=messages.ERROR)
                if len(errors) > 10:
                    self.message_user(request, f"Ещё ошибок: {len(errors) - 10}", level=messages.ERROR)
                self.message_user(
                    request,
                    f"Импорт завершён: создано {imported_count}, обновлено {updated_count}, ошибок {len(errors)}.",
                )
                return redirect("..")
        else:
            form = ProductImportForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "form": form,
            "title": "Импорт товаров",
        }
        return render(request, "admin/shop/product/import_products.html", context)

    def _import_products_from_file(self, upload, as_csv: bool):
        rows = self._read_csv_rows(upload) if as_csv else self._read_xlsx_rows(upload)
        imported_count = 0
        updated_count = 0
        errors: list[str] = []

        for idx, raw_row in enumerate(rows, start=2):
            row = {key: (raw_row.get(key) or "").strip() for key in PRODUCT_IMPORT_EXPORT_FIELDS}
            for legacy_key in ("is_best_price", "is_closed_sale", "is_new"):
                if legacy_key in raw_row:
                    row[legacy_key] = (raw_row.get(legacy_key) or "").strip()
            try:
                sku = row["sku"]
                if not sku:
                    raise ValueError("SKU обязателен")
                category = Category.objects.filter(slug=row["category_slug"]).first()
                if not category:
                    raise ValueError(f"Категория '{row['category_slug']}' не найдена")

                defaults = {
                    "category": category,
                    "name": row["name"] or sku,
                    "slug": row["slug"] or sku,
                    "price": self._parse_decimal(row["price"], "price"),
                    "old_price": self._parse_decimal(row["old_price"], "old_price", allow_blank=True),
                    "image_url": row["image_url"],
                    "description": row["description"],
                    "stock_store": self._parse_int(row["stock_store"], "stock_store"),
                    "stock_warehouse": self._parse_int(row["stock_warehouse"], "stock_warehouse"),
                }
                label_slugs = self._parse_label_slugs(row)
                obj = Product.objects.filter(sku=sku).first()
                created = obj is None
                if created:
                    obj = Product(sku=sku, **defaults)
                else:
                    for field, value in defaults.items():
                        setattr(obj, field, value)
                obj.full_clean()
                obj.save()
                obj.labels.set(self._labels_for_slugs(label_slugs))
                if created:
                    imported_count += 1
                else:
                    updated_count += 1
            except Exception as exc:
                errors.append(f"Строка {idx}: {exc}")
        Category.sync_product_counts()
        return imported_count, updated_count, errors

    def _read_csv_rows(self, upload):
        content = upload.read().decode("utf-8-sig")
        reader = csv.DictReader(content.splitlines(), delimiter=";")
        return list(reader)

    def _read_xlsx_rows(self, upload):
        wb = load_workbook(filename=upload, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data = []
        for row in rows[1:]:
            data.append(
                {
                    headers[i]: str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                    for i in range(len(headers))
                },
            )
        return data

    @staticmethod
    def _parse_decimal(raw: str, field_name: str, allow_blank: bool = False):
        if not raw and allow_blank:
            return None
        value = (raw or "").replace(",", ".").strip()
        try:
            return Decimal(value)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError(f"Некорректное значение {field_name}: '{raw}'") from exc

    @staticmethod
    def _parse_int(raw: str, field_name: str):
        value = (raw or "").strip()
        if not value:
            return 0
        try:
            return int(value)
        except ValueError as exc:
            raise ValueError(f"Некорректное значение {field_name}: '{raw}'") from exc

    @staticmethod
    def _parse_bool(raw: str):
        return str(raw).strip().lower() in {"1", "true", "yes", "да", "on"}

    def _parse_label_slugs(self, row: dict) -> list[str]:
        raw = (row.get("labels") or "").strip()
        if raw:
            return [part.strip() for part in raw.replace(",", ";").split(";") if part.strip()]
        legacy_map = (
            ("is_best_price", "best-price"),
            ("is_closed_sale", "closed-sale"),
            ("is_new", "new"),
        )
        slugs: list[str] = []
        for field, slug in legacy_map:
            if field in row and self._parse_bool(row[field]):
                slugs.append(slug)
        return slugs

    @staticmethod
    def _labels_for_slugs(slugs: list[str]):
        if not slugs:
            return []
        labels = list(ProductLabel.objects.filter(slug__in=slugs))
        found = {label.slug for label in labels}
        missing = [slug for slug in slugs if slug not in found]
        if missing:
            raise ValueError(f"Метки не найдены: {', '.join(missing)}")
        order = {slug: idx for idx, slug in enumerate(slugs)}
        return sorted(labels, key=lambda label: order.get(label.slug, 999))


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ("product", "quantity", "unit_price")


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ("id", "customer_name", "phone", "total_price", "status", "created_at")
    list_filter = ("status", "created_at")
    search_fields = ("customer_name", "phone", "email")
    inlines = [OrderItemInline]


@admin.register(Favorite)
class FavoriteAdmin(admin.ModelAdmin):
    list_display = ("user", "product", "created_at")
    list_filter = ("created_at",)
    search_fields = ("user__email", "user__username", "product__name", "product__sku")


@admin.register(UserProfile)
class UserProfileAdmin(admin.ModelAdmin):
    list_display = ("user", "phone", "delivery_address")
    search_fields = ("user__email", "user__username", "phone", "delivery_address")


@admin.register(Organization)
class OrganizationAdmin(admin.ModelAdmin):
    list_display = ("name", "inn", "kpp", "owner", "created_at")
    list_filter = ("created_at",)
    search_fields = ("name", "inn", "kpp", "owner__email", "owner__username")


class PublishedContentAdmin(admin.ModelAdmin):
    list_display = ("title", "slug", "start_date", "end_date", "is_published", "published_at")
    list_filter = ("is_published", "published_at", "start_date")
    search_fields = ("title", "slug", "excerpt", "body")
    prepopulated_fields = {"slug": ("title",)}
    date_hierarchy = "published_at"
    ordering = ("-published_at", "-id")
    fieldsets = (
        (None, {"fields": ("title", "slug", "excerpt", "image_url", "start_date", "end_date", "is_published", "published_at")}),
        ("Текст", {"fields": ("body",)}),
    )


@admin.register(Article)
class ArticleAdmin(PublishedContentAdmin):
    pass


@admin.register(News)
class NewsAdmin(PublishedContentAdmin):
    pass


@admin.register(Promotion)
class PromotionAdmin(PublishedContentAdmin):
    pass
