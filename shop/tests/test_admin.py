import csv
from io import BytesIO

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from shop.admin import ProductAdmin
from shop.models import Product
from shop.models import ProductLabel
from shop.tests.factories import (
    make_category,
    make_product,
    make_product_spec_value,
    make_spec_attribute,
)


class ProductAdminImportExportTests(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.admin = ProductAdmin(Product, self.site)
        self.factory = RequestFactory()
        self.category = make_category(name="Метизы", slug="metizy-admin")
        self.product = make_product(
            category=self.category,
            name="Гайка M8",
            slug="gaika-m8",
            sku="SKU-A-001",
        )
        attr = make_spec_attribute(category=self.category, name="DIN", sort_order=1)
        make_product_spec_value(product=self.product, attribute=attr, value="934")

    def test_admin_change_form_has_preview_button(self):
        user = get_user_model().objects.create_superuser("admin", "admin@test.local", "pass")
        self.client.force_login(user)
        url = reverse("admin:shop_product_change", args=[self.product.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Просмотр")
        self.assertContains(response, f'/product/{self.product.slug}/"')
        self.assertContains(response, 'target="_blank"')

    def test_admin_export_csv_contains_spec_columns(self):
        request = self.factory.post("/admin/shop/product/", {"use_csv": "on"})
        response = self.admin.export_products(request, Product.objects.all())
        self.assertEqual(response.status_code, 200)
        content = response.content.decode("utf-8-sig")
        rows = list(csv.reader(content.splitlines(), delimiter=";"))
        header = rows[0]
        self.assertIn("spec__DIN", header)
        self.assertEqual(rows[1][header.index("spec__DIN")], "934")

    def test_admin_export_xlsx_contains_spec_columns(self):
        request = self.factory.post("/admin/shop/product/", {})
        response = self.admin.export_products(request, Product.objects.all())
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        header = [c.value for c in ws[1]]
        self.assertIn("spec__DIN", header)
        self.assertEqual(ws.cell(row=2, column=header.index("spec__DIN") + 1).value, "934")

    def test_admin_import_csv_creates_product(self):
        csv_payload = (
            "category_slug;name;slug;sku;price;old_price;image_url;description;stock_store;stock_warehouse;labels\n"
            "metizy-admin;Болт M10;bolt-m10;SKU-A-777;210;310;https://example.com/new.png;Импорт;12;33;new\n"
        )
        ProductLabel.objects.get(slug="new")
        upload = SimpleUploadedFile("products.csv", csv_payload.encode("utf-8"), content_type="text/csv")
        created, updated, errors = self.admin._import_products_from_file(upload=upload, as_csv=True)
        self.assertEqual(created, 1)
        self.assertEqual(updated, 0)
        self.assertEqual(errors, [])
        imported = Product.objects.get(sku="SKU-A-777", slug="bolt-m10")
        self.assertEqual(list(imported.labels.values_list("slug", flat=True)), ["new"])
